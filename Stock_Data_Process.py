import random
import pandas as pd
import codecs
import csv
import openpyxl
from openpyxl import *
import os
import matplotlib.pyplot as plt
from datetime import datetime
from openpyxl.utils import column_index_from_string
from pandas import DataFrame
from xlrd import xldate_as_tuple
import time

def addStock_data(path, dirpath):
    data = pd.read_pickle(path)
    data = pd.DataFrame(data)
    #print(data.columns.values.tolist())
    #data = data.drop('Unnamed: 0', axis=1)
    for x in data.groupby('代码'):
        if not os.path.exists(dirpath):
            os.mkdir(dirpath)
        if not os.path.exists(dirpath + "/" + str(x[0]) + ".csv"):
            df = pd.DataFrame(x[1])
            df.reset_index(drop=True, inplace=True)
            df.to_csv(dirpath + "/" + str(x[0]) + ".csv", encoding='utf-8')
        else:
            past_data = pd.read_csv(dirpath + "/" + str(x[0]) + ".csv")
            past_data = past_data.drop('Unnamed: 0', axis=1)
            past_data = pd.DataFrame(past_data)
            new_df = pd.DataFrame(x[1])
            past_data = past_data.append(new_df, ignore_index=True)
            past_data.reset_index(drop=True, inplace=True)
            past_data.to_csv(dirpath + "/" + str(x[0]) + ".csv", encoding='utf-8')


# 将sheet里所有在checklist里的列复制到new_sheet中
def removeUselessCol(sheet, new_sheet, check_list):
    rows = sheet.max_row
    cols = sheet.max_column
    k = 1
    for i in range(1, cols + 1):
        if sheet.cell(row=1, column=i).value in check_list:
            # print(ws.cell(row=1, column=i).value)
            j = 1
            for cell in list(sheet.columns)[i - 1]:
                new_sheet.cell(row=j, column=k).value = cell.value
                if i == 1 and j > 1:
                    new_sheet.cell(row=j, column=k).value = new_sheet.cell(row=j, column=k).value.strftime('%Y-%m-%d')
                #修改代码列
                if i == 2 and j >= 2 and type(new_sheet.cell(j, k).value) == str:
                    new_sheet.cell(j, k).value = new_sheet.cell(j, k).value.strip('=')
                    new_sheet.cell(j, k).value = new_sheet.cell(j, k).value.strip('"')
                elif i == 2 and j >= 2 and type(new_sheet.cell(j, k).value) == int:
                    new_str = str(new_sheet.cell(j, k).value)
                    while len(new_str) < 6:
                        new_str = "0" + new_str
                    new_sheet.cell(j, k).value = new_str
                j += 1
            k += 1
        else:
            continue
    return


# 将checklist中的都进行排序，并添加rank
def addRank_Stock(path):
    df = pd.read_excel(path, converters={u'代码': str}) #, engine='openpyxl'
    df = pd.DataFrame(df)
    new_check_list = df.columns.values.tolist()
    for name in new_check_list[3:]:
        new_str = name + 'Rank'
        df[new_str] = df[name].rank(method='min', ascending=False, numeric_only=True)
    df.to_pickle('./Data/test.pkl')
    df.to_csv('./Data_Copy/test.csv', encoding='utf-8')


#检查sheet某一列是否含有某个值
def is_exist(value, sheet, index):
    for cell in list(sheet.columns)[index]:
        if value == str(cell.value):
            # print("存在非法字符，仍需删除")
            return True
        # else:
        #     print("不存在")


# 删除sheet中'换手Z'为空的所有行
def deleteInvalidStr(sheet):
    while is_exist('--  ', sheet, 3):
        for row in sheet.iter_rows():
            for cel in row:
                if sheet.cell(1, cel.column).value == '换手Z':# and (cel.value == '--  ' or cel.value == 0) and :
                    if type(cel.value) == str and cel.value == '--  ':
                        row_number = row[1].row
                        sheet.delete_rows(row_number)
                        print("删除成功")


def replaceInvalidValue():
    return


# 获取当前目录下的CSV文件名
def name(path):
    file_name = []
    # 将当前目录下的所有文件名称读取进来
    a = os.listdir(path)
    a.sort()
    for j in a:
        # 判断是否为xlsx文件，如果是则存储到列表中
        if os.path.splitext(j)[1] == '.xlsx':
            file_name.append(j)
    return file_name

def stock_dir_process(path, dirpath):
    file_name = name(path)
    for fn in file_name:
        single_stock_process(path + fn, dirpath)
    return


def single_stock_static(filename, operation, num):
    plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
    # 画图
    # 如何创建同y轴的上下三联图，需要plt.subplots()
    fig, axes = plt.subplots(nrows=len(operation), ncols=1, sharex='all', figsize=(30, 15), squeeze=False)
    fig.suptitle("股票统计图", fontsize=20)
    ll = []
    x = 0
    for fn in filename:
        for opt in operation:
            file = pd.read_csv('./Stock_Data/' + fn + '.csv')
            df = pd.DataFrame(file)
            ylist = df[opt].to_list()
            xlist = df['日期'].to_list()
            ll = xlist
            # subline1
            axes1 = axes[x][0]
            axes1.plot(xlist, ylist, marker='o', color='black', label=fn + ',' + opt)
            axes1.grid(which='major', axis='both', linestyle='--')

            # subline2
            df['Ma5'] = df[opt].rolling(window=5).mean()
            axes1.plot(xlist, df['Ma5'], color='yellow', label='Ma5')

            # subline3
            df['Ma20'] = df[opt].rolling(window=20).mean()
            axes1.plot(xlist, df['Ma20'], color='red', label='Ma20')

            # subline4
            strr = 'Ma' + num
            n = int(num)
            df[strr] = df[opt].rolling(window=n).mean()
            axes1.plot(xlist, df[strr], color='blue', label=strr)
            axes1.legend(fontsize = 20)
            x = x + 1
        axes[x-1][0].set_xticks(ll)
        axes[x-1][0].set_xticklabels(labels=ll, rotation=90)
        fig.savefig('./Stock_Graph/stock.png')
    return


def randomcolor():
    colorArr = ['1','2','3','4','5','6','7','8','9','A','B','C','D','E','F']
    color = ""
    for i in range(6):
        color += colorArr[random.randint(0,14)]
    return "#" + color


def single_stock_process(path, dirpath):
    '''
    Stage 1: DATA CLEANING
    Remove all useless columns
    Remove all invalid strings
    '''
    sTime = time.time()
    check_list = ['日期', '代码', '名称', '竞价量', '涨幅%', '开买量', '开卖量', '买量', '卖量', '开盘', '最高', '最低', '收盘', '总量',
                  '量比', '振幅%', '总金额', '开盘金额', '内盘', '外盘', '内外比', '攻击波%', '回头波%', '强弱度%', '活跃度', '笔均量', '换手Z',
                  '当日___净流入', '净买率%', '当日___相对流量%', '当日___超大单', '当日___大单', '当日___中单', '当日___小单',
                  'BS', 'WIN', 'RSI1', '今日强', '3日强', '5日强', '10日强', '20日强', '60日强', '一年来', '年初至今']
    wb = openpyxl.load_workbook(path)
    ws = wb.worksheets[0]
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    # 去除没用列
    removeUselessCol(ws, new_ws, check_list)
    # 去除非法string
    deleteInvalidStr(new_ws)
    new_wb.save(filename='./Data/test.xlsx')
    # 添加排名
    addRank_Stock('./Data/test.xlsx')
    # 关闭文件
    wb.close()
    new_wb.close()
    os.remove('./Data/test.xlsx')
    '''
    Stage 2: Add data to stock file saperately
    '''
    addStock_data('./Data/test.pkl', dirpath)
    os.remove('./Data/test.pkl')
    eTime = time.time()
    print("单次文件操作时间：", eTime - sTime)
    return

#检查输入是否为空
def checkInput(inp):
    if inp == "":
        return True
    else:
        return False


# 条件选股
def stockSelect(file, dirpath, selectName, number):
    df = pd.read_csv(file)
    df = pd.DataFrame(df)
    selectName = list(map(str, selectName.split()))
    outputfile = df
    # outputfile = df[(df['买量Rank'] <= int(number)) & (df['卖量Rank'] <= int(number))]
    for name in selectName:
        outputfile = outputfile[(outputfile[str(name)] <= int(number))]
    outputfile = outputfile.drop('Unnamed: 0', axis=1)
    outputfile.reset_index(drop=True, inplace=True)
    outputfile.to_csv(dirpath + "/day.csv", encoding='gbk')
    return
#./Data/test.xlsx     ./Stock_Select      量比 换手Z 涨幅%
# 量比Rank 涨幅%Rank 笔均量Rank


def Add_or_Delete(opt, path, nrow):
    file_name = []
    if opt == "1":
        GetName(file_name, path)
        for name in file_name:
            df = pd.read_csv(path + "/" + name, index_col=0)
            if int(nrow) in df.index:
                df = df.drop(df.index[int(nrow)])
                df.reset_index(drop=True, inplace=True)
                df.to_csv(path + "/" + name, encoding='utf-8-sig')
    return


#获取当前目录下的CSV文件名
def GetName(file_name, path):
    #将当前目录下的所有文件名称读取进来
    a = os.listdir(path)
    a.sort()
    for j in a:
        #判断是否为CSV文件，如果是则存储到列表中
        if os.path.splitext(j)[1] == '.csv':
            file_name.append(j)


if __name__ == '__main__':
    while True:
        print("股票数据操作")
        print("1.单日数据添加以及处理(排序操作待添加)")
        print("2.多日数据添加以及处理(有待于操作1整合)")
        print("3.股票走势统计与对比")
        print("4.单日股票条件选股")
        print("5.股票删除或添加一行")
        print("6.退出")
        oprator = input("请输入所需操作：")
        while  oprator == "" or (int(oprator) < 1 or int(oprator) > 6):
            print("请输入正确操作数字")
            oprator = input("请输入所需操作：")
        if oprator == "1":
            while True:
                path1 = input("请输入需要处理的文件目录：")
                dirpath = input("请输入存放股票代码目录(如不存在，将自动创建该目录)：")
                if not os.path.isfile(path1):
                    print("文件不存在，请重新输入！")
                    continue
                else:
                    break
            startTime = time.time()
            print("正在操作......")
            single_stock_process(path1, dirpath)
            endTime = time.time()
            print("操作成功！" + "本次操作时长：" + str(endTime - startTime))
            print("\n\n" + "<------------------------------------------------------------->")
            continue
        elif oprator == "2":
            while True:
                path2 = input("请输入需要处理的文件夹目录：")
                if not os.path.isdir(path2):
                    print("文件夹目录不存在！，请重新输入")
                    continue
                else:
                    break
            dirpath2 = input("请输入存放股票代码目录(如不存在，将自动创建该目录)：")
            st = time.time()
            print("操作开始.....")
            stock_dir_process(path2, dirpath2)
            et = time.time()
            print("操作成功！操作时间：" + str(et - st))
            print("\n\n" + "<------------------------------------------------------------->")
            continue
        elif oprator == "3":
            while True:
                file = input("请输入需要处理的股票代码 (输入方式举例：000001 000002......, 若返回上一层请输入exit)：")
                if file == "":
                    print("输入不能为空")
                    continue
                elif file == "exit":
                    break
                filename = list(map(str, file.split()))
                while True:
                    opt = input("请输入查看的统计方式 (输入方式举例：量比 换手Z......，若返回上一层请输入exit)：")
                    if opt == "exit":
                        break
                    elif opt == "":
                        print("输入不能为空")
                        continue
                    else:
                        operation = list(map(str, opt.split()))
                        num = input("请输入想查看的均线(输入举例: 5，10，20.....，默认已含有5日和20日均线): ")
                        single_stock_static(filename, operation, num)
                        print("操作成功！")
            print("\n" + "<------------------------------------------------------------->")
            continue
        elif oprator == "4":
            while True:
                file = input("请输入需要处理的文件目录(举例：./Data_Copy/test.csv)：")
                while checkInput(file):
                    file = input("请输入需要处理的文件目录：")
                dirpath = input("请输入存放选股数据的目录(如不存在，将自动创建该目录，举例：./Stock_Select)：")
                while checkInput(dirpath):
                    dirpath = input("请输入存放选股数据的目录(如不存在，将自动创建该目录, 举例：./Stock_Select)：")
                selectName = input("请输入需要查询的条件(输入方式举例：量比 换手Z 涨幅%.....)：")
                while checkInput(selectName):
                    selectName = input("请输入需要查询的条件(输入方式举例：量比 换手Z 涨幅%.....)：")
                number = input("请输入需要筛选排名条件：")
                while checkInput(number):
                    number = input("请输入需要筛选排名条件：")
                if not os.path.isfile(file):
                    print("文件不存在，请重新输入！")
                    continue
                else:
                    break
            startTime = time.time()
            print("正在筛选......")
            stockSelect(file, dirpath, selectName, number)
            endTime = time.time()
            print("操作成功！" + "本次操作时长：" + str(endTime - startTime))
            print("\n" + "<------------------------------------------------------------->")
            continue
        elif oprator == "5":
            opt = input("请输入需要的操作(输入1代表删除，输入2代表添加)：")
            path = input("请输入要操作的目录：")
            nrow = input("请输入要操作的行数：")
            print("正在修改......")
            startTime = time.time()
            Add_or_Delete(opt, path, nrow)
            endTime = time.time()
            print("操作成功！" + "本次操作时长：" + str(endTime - startTime))
            print("\n" + "<------------------------------------------------------------->")
        else:
            break

